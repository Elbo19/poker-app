#!/usr/bin/env python3
import openpyxl
import requests
import json
import re

# API endpoint
API_URL = "http://localhost:8080/api/compare"

def parse_cards(card_string):
    """Parse card string into list of card codes."""
    if not card_string or card_string == 'None':
        return []
    
    # Handle special case where cards don't matter (represented by – or -)
    if card_string.strip() in ['–', '-', '—']:
        return None  # Special marker for "cards don't matter"
    
    # Remove extra whitespace and non-breaking spaces
    card_string = card_string.replace('\xa0', ' ').strip()
    
    # Split by spaces and filter empty strings
    cards = [c.strip() for c in card_string.split() if c.strip()]
    
    return cards

def test_hand_comparison(community_cards, player1_cards, player2_cards, expected_result, row_num):
    """Test a single hand comparison."""
    
    # Parse cards
    p1_hole = parse_cards(player1_cards)
    p2_hole = parse_cards(player2_cards)
    comm = parse_cards(community_cards)
    
    # Skip only if we truly have no data
    if not comm:
        return None, "Skipped (no data)"
    
    # Handle special case where hole cards don't matter (e.g., Royal Flush in community)
    # Use arbitrary cards that aren't in the community
    if p1_hole is None or p2_hole is None:
        # Get cards not in community
        all_cards = []
        for suit in ['H', 'D', 'C', 'S']:
            for rank in ['2', '3', '4', '5', '6', '7', '8', '9', 'T', 'J', 'Q', 'K', 'A']:
                all_cards.append(suit + rank)
        
        # Remove community cards
        available = [c for c in all_cards if c not in comm]
        
        # Assign arbitrary unique hole cards
        if p1_hole is None:
            p1_hole = available[:2]
            available = available[2:]
        if p2_hole is None:
            p2_hole = available[:2]
    
    # Skip if we still don't have valid cards
    if not p1_hole or not p2_hole:
        return None, "Skipped (invalid cards)"
    
    # Prepare request
    payload = {
        "player1HoleCards": p1_hole,
        "player1CommunityCards": comm,
        "player2HoleCards": p2_hole,
        "player2CommunityCards": comm
    }
    
    try:
        # Make API request
        response = requests.post(API_URL, json=payload, timeout=5)
        
        if response.status_code != 200:
            return False, f"API error: {response.status_code} - {response.text}"
        
        result = response.json()
        
        if not result.get('success', False):
            return False, f"API returned error: {result.get('error', 'Unknown error')}"
        
        # Determine expected winner
        if "hand 1 > hand 2" in expected_result or "hand 1 >hand 2" in expected_result or "hand 1> hand 2" in expected_result:
            expected_winner = "player1"
        elif "hand 2 > hand 1" in expected_result or "hand 2 >hand 1" in expected_result or "hand 2> hand 1" in expected_result:
            expected_winner = "player2"
        elif "hand 1 < hand 2" in expected_result or "hand 1 <hand 2" in expected_result or "hand 1< hand 2" in expected_result:
            expected_winner = "player2"
        elif "hand 2 < hand 1" in expected_result or "hand 2 <hand 1" in expected_result or "hand 2< hand 1" in expected_result:
            expected_winner = "player1"
        elif "hand 1 = hand 2" in expected_result or "hand 1 =hand 2" in expected_result or "hand 1= hand 2" in expected_result:
            expected_winner = "tie"
        else:
            return None, f"Unknown expected result format: {expected_result}"
        
        actual_winner = result.get('winner', '')
        
        if actual_winner == expected_winner:
            return True, f"✓ {result['player1Hand']} vs {result['player2Hand']} = {actual_winner}"
        else:
            return False, f"✗ Expected {expected_winner}, got {actual_winner}. P1: {result['player1Hand']}, P2: {result['player2Hand']}"
    
    except requests.exceptions.ConnectionError:
        return False, "Cannot connect to API server. Is it running on port 8080?"
    except Exception as e:
        return False, f"Error: {str(e)}"

def main():
    # Load Excel file
    wb = openpyxl.load_workbook('/Users/elbetel/projects/poker-app/Texas HoldEm Hand comparison test cases.xlsx')
    ws = wb.active
    
    print("="*80)
    print("TESTING POKER HAND COMPARISON")
    print("="*80)
    
    passed = 0
    failed = 0
    skipped = 0
    failures = []
    
    # Track previous row for permutations
    prev_community = None
    prev_player1 = None
    prev_player2 = None
    
    # Process each row (skip header)
    for row_num in range(2, ws.max_row + 1):
        row = ws[row_num]
        
        hand_type = row[0].value
        community_cards = row[1].value
        player1_cards = row[2].value
        player1_hand = row[3].value
        player2_cards = row[4].value
        player2_hand = row[5].value
        expected_result = row[6].value
        comment = row[8].value if len(row) > 8 else None
        
        # Skip empty rows (no expected result)
        if not expected_result or expected_result.strip() == "":
            skipped += 1
            continue
        
        # For permutation rows (where cards are None), use previous row's cards
        if community_cards is None and prev_community is not None:
            community_cards = prev_community
        if player1_cards is None and prev_player1 is not None:
            player1_cards = prev_player1
        if player2_cards is None and prev_player2 is not None:
            player2_cards = prev_player2
        
        # Update previous row tracking
        if community_cards is not None:
            prev_community = community_cards
        if player1_cards is not None:
            prev_player1 = player1_cards
        if player2_cards is not None:
            prev_player2 = player2_cards
        
        result, message = test_hand_comparison(
            community_cards, 
            player1_cards, 
            player2_cards, 
            expected_result, 
            row_num
        )
        
        if result is None:
            skipped += 1
            continue
        elif result:
            passed += 1
            print(f"Row {row_num:2d} [{hand_type:20s}]: {message}")
        else:
            failed += 1
            failure_info = {
                'row': row_num,
                'hand_type': hand_type,
                'community': community_cards,
                'player1': player1_cards,
                'player2': player2_cards,
                'expected': expected_result,
                'comment': comment,
                'message': message
            }
            failures.append(failure_info)
            print(f"Row {row_num:2d} [{hand_type:20s}]: {message}")
    
    print("\n" + "="*80)
    print(f"RESULTS: {passed} passed, {failed} failed, {skipped} skipped")
    print(f"Total data rows (excluding header): {ws.max_row - 1}")
    print(f"Valid test rows: {passed}")
    print(f"Empty separator rows: {skipped}")
    print("="*80)
    
    if failures:
        print("\nFAILED TEST CASES:")
        print("="*80)
        for f in failures:
            print(f"\nRow {f['row']}: {f['hand_type']}")
            print(f"  Community: {f['community']}")
            print(f"  Player 1:  {f['player1']}")
            print(f"  Player 2:  {f['player2']}")
            print(f"  Expected:  {f['expected']}")
            if f['comment']:
                print(f"  Comment:   {f['comment']}")
            print(f"  Result:    {f['message']}")
    
    return failed == 0

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
